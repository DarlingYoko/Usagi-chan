import discord, pytz
from discord.ext import commands, tasks
from bin.converters import *
from openpyxl import load_workbook
from random import choice
from os import listdir
from discord import File
from cogs.birthday.utils import *
from datetime import datetime





class Birthday(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
        self.config = bot.config
        self.birthdays_notify.start()

    @tasks.loop(minutes=1)
    async def birthdays_notify(self):
        timezone = pytz.timezone("Europe/Moscow")
        time = datetime.now(timezone)
        if time.hour == 11 and time.minute == 25:


            download = downloadShedule()
            if not download:
                print("Download failed")
            wb = load_workbook('cogs/birthday/birthdays.xlsx')
            path = './files/photo/birthday/'
            sheet = wb['Лист1']

            channel = await self.bot.fetch_channel(self.config['channel']['main'])



            row_count = 5
            column_count = 2

            birthday_names = []

            while column_count != 14:
                name = sheet.cell(row = row_count, column = column_count).value
                while name:
                    name = name.split('-')
                    day = name[0].strip()
                    # print(name[1].split(','))
                    id = name[1].split(',')[2].strip()
                    name = name[1].split(',')[0].strip()
                    birthday_names.append({'day': int(day), 'month': column_count - 1, 'name': name, 'uid': id})
                    row_count += 1
                    name = sheet.cell(row = row_count, column = column_count).value
                column_count += 1
                row_count = 5


            for person in birthday_names:
                # print(time.day, person['day'], time.month, person['month'])
                if time.day == person['day'] and time.month == person['month']:
                    texts = [
                        f'Слышь, <@{person["uid"]}>, позырил твои фотки, ты прям кайфовая. Ну тип глаза ровно смотрят не косые, нос не до подбородка, зубы вроде все на месте, прыщей на лице нет, туловище тоже в хорошом состояний. Ноги ходят, руки держат все как надо, збс прям. Ты норм. Ты молодец. Дай бог тебе здоровья.\n С днём рожденья!',
                        f'<@{person["uid"]}> фантастически красивая, я просто хуею с твоих черт лица и особенно улыбкаааа... я слежу за тобой с самого начала, ну не прям так слежу, ну короч не суть, вот так смотрю на тебя и мне кажется , что ты как будто не в ту жизненную тарелку попала, ты как будто в заперти и боишься выйти, кажется ты должна ходить по клубам бухать гулять и тд , прости если хуйню несу , я вообще хз почему решил тебе написать ну пох!Ну ты прям сосочная , делай больше фоток и улыбайся тебе очень идёт , правда !\n С днём рожденья!',
                        f'С днюхой тебя, <@{person["uid"]}>! С сорока с чем то летием!',
                        f'Сегодня особенный день для <@{person["uid"]}>! Поздравляю с Днём варенья - желаю успехов в любом деле, радости и процветания!',
                        f'Привет, <@{person["uid"]}>! Сегодня твой День рожденья - самое время провести его в отличной компании. Такое солнышко как ты сделает любой день лучше!',
                        f'<@{person["uid"]}>\n🌹🍒Мило🍒🌹 🌸💙Ярко💙🌸 💖😙Красотка😙 🌿💐Сногшебательна💐🌿 😍😅😅😍 😎😏Круто😏😎 ✨👑Велеколепно👑✨ 🎀💕ТОП🎀💕 🍬🍭Конфетка🍭🍬 😝😋Няшно😋😝 🌺🍃По королевски🍃🌺 🌾🌻Ослепительно🌻🌾 🐰🌹Зайка🌹🐰 🐱🌵Милафф🌵🐱 🐳🐋Не струя,а фонтан🐳🐋\n С днём рожденья!',
                        f'<@{person["uid"]}>\n, так чего я пишу то, с др тебя. Шоб cock стоял и деньги были<:HmmFlower:861241411805642782>',
                        ]
                    photo = path + choice(listdir(path))
                    print(f'Сегодня др у {person["name"]}, {photo}')
                    file = File(photo)
                    text = choice(texts)
                    await channel.send(content=text, file=file)



def setup(bot):
    bot.add_cog(Birthday(bot))
