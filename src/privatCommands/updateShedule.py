from openpyxl import load_workbook
from google.downloadFiles import downloadShedule
from src.functions import createEmbed
from datetime import datetime, timedelta
from time import mktime

async def checkNotification(self):
    requestsList = self.db.getData()
    for (userId, data, users) in requestsList:
        time = datetime.fromtimestamp(float(data))
        users = eval(users)
        #print(userId, data, users)
        try:
            if time - datetime.now() - timedelta(hours = 1) <= timedelta(minutes = 10):
                answer = ''
                for user in users:
                    answer += '<@{0}> '.format(user)
                answer += '\nСеанс начнётся через 10 минут!'
                channelShedule = await self.client.fetch_channel(self.config['sheduleData']['sheduleChannel'])
                channelChat = await self.client.fetch_channel(791014331323777044)

                message = await channelShedule.fetch_message(829293778267406357)
                embed = message.embeds[0]
                await channelChat.send(answer, embed = embed)
                self.db.remove(userId, 'SHEDULE')
        except Exception as e:
            print(e)



def getRussainDay(day):
    days = {
            'Monday': 'Понедельник',
            'Tuesday': 'Вторник',
            'Wednesday': 'Среда',
            'Thursday': 'Четверг',
            'Friday': 'Пятница',
            'Saturday': 'Суббота',
            'Sunday': 'Воскресенье'
            }
    if day in days.keys():
        return days[day]
    return 0

def getRussainMonth(month):
    months = {
                'January':'Января',
                'February': 'Февраля',
                'March': 'Марта',
                'April': 'Апреля',
                'May': 'Мая',
                'June': 'Июня',
                'July': 'Июля',
                'August': 'Августа',
                'September': 'Сентября',
                'October': 'Октября',
                'November': 'Ноября',
                'December': 'Декабря'
    }
    if month in months.keys():
        return months[month]
    return 0

def updateShedule():

    if not downloadShedule():
        return 0

    wb = load_workbook('D:\Projects\Discord\Yoko-bot\\google\\shedule.xlsx')
    embeds = {}

    sheet = wb['sheduleList']

    countCell = 4

    day = sheet.cell(row = countCell, column = 2).value

    while day:
        streamerCell = countCell + 1
        streamer = sheet.cell(row = streamerCell, column = 4).value
        dayNumber = sheet.cell(row = countCell, column = 2).value.strftime('%A %d %B').split(' ')

        months = [
                    'January',
                    'February',
                    'March',
                    'April',
                    'May',
                    'June',
                    'July',
                    'August',
                    'September',
                    'October',
                    'November',
                    'December']


        monthCount = months.index(dayNumber[2]) + 1
        if monthCount < 10:
            monthCount = '0' + str(monthCount)
        monthCount = str(monthCount)


        dayName, dayCount, monthName = getRussainDay(dayNumber[0]), int(dayNumber[1]), getRussainMonth(dayNumber[2])

        fieldName = '<:Primogem:828706824370126920> {0} - *{1} {2}*'.format(dayName, dayCount, monthName)
        fieldValue = ''
        while streamer:
            titleName = sheet.cell(row = streamerCell, column = 5).value
            series = sheet.cell(row = streamerCell, column = 6).value
            time = sheet.cell(row = streamerCell, column = 7).value
            hall = sheet.cell(row = streamerCell, column = 8).value
            sound = sheet.cell(row = streamerCell, column = 9).value
            fieldValue = '<:Keqing:771095540456489020> Стримит: <@{0}> **{1}** ({2}) - {3} - **Зал {4}** - {5}\n'.format(int(streamer), titleName, series, time.strftime('%H:%M'), int(hall), sound)
            streamerCell += 1
            streamer = sheet.cell(row = streamerCell, column = 4).value


            date = '2021-{0}-{1} {2}'.format(monthCount, dayNumber[1], time)
            format = "%Y-%m-%d %H:%M:%S"
            datetime_obj = datetime.strptime(date, format)
            seconds = mktime(datetime_obj.timetuple())


            embeds[createEmbed(title = fieldName, description = fieldValue)] = seconds



        countCell = 1 + streamerCell
        day = sheet.cell(row = countCell, column = 2).value


    return embeds
